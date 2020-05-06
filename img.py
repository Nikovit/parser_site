from bs4 import BeautifulSoup
import requests
import re
import openpyxl
import urllib.request

# URL страницы поиска
url_src = 'https://www.sds-group.ru/search.htm?search='

# Файл артикулов
file = "sds-group.xlsx"

# Инициализируем библиотеку для работы с Excel и считываем значения колонок
wb_obj = openpyxl.load_workbook(file)
sheet_obj = wb_obj.active
m_row = sheet_obj.max_row

for i in range(1, m_row + 1):
    # Считываем значение ячейки
    cell_obj = sheet_obj.cell(row=i, column=1)
    print('-------------------------')
    print('Артикул: ' + str(cell_obj.value))
    # Формируем URL с поиском переменная url_src + значение ячейки
    url = url_src + str(cell_obj.value)
    # Делаем http запрос библиотекой requests и помещаем ответ сервера в переменную r
    r = requests.get(url)
    # Инициализируем парсер BeautifulSoup
    soup = BeautifulSoup(r.text, 'html.parser')
    # Находим на полученной странице с сайта элемент с классами flex-3 m-flex t-left(именнно в этом элементе содержится ссылка на детальную страницу товара)
    url_item = soup.find_all(class_='flex-3 m-flex t-left')
    url_item_str = str(url_item)
    item = re.findall('href=[\'"]?([^\'" >]+)', url_item_str)
    # Проверяем нашелся ли вообще нужный артикул на сайте? если нет то переходим к следующему артиклу из файла
    try:
        item_href = str(item[0])
        print('https://www.sds-group.ru' + item_href)
    except IndexError:
        print('Артикул на сайте не найден')
        continue

    # Если артикул нашелся то переходим на страницу детального просмотра
    r = requests.get('https://www.sds-group.ru' + item_href)
    soup2 = BeautifulSoup(r.text, 'html.parser')

    # Изображение находится в div с классом product-img, находим его и получаем прямую ссылку на картинку
    img_item = soup2.find("div", {"class": "product-img"})
    img = img_item.img['src']
    img_src = 'https://www.sds-group.ru' + img

    print(img_src)

    # Скачиваем файл в папку img
    try:
        urllib.request.urlretrieve(img_src, './img/' + str(cell_obj.value) + '.jpg')
    except:
        print('Не получилось сохранить файл')
