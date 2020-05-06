from bs4 import BeautifulSoup
import requests
import re
import openpyxl

# URL страницы поиска
url_src = 'https://www.sds-group.ru/search.htm?search='

# Файл артикулов
file = "rexant_txt.xlsx"

wb_obj = openpyxl.load_workbook(file)
sheet_obj = wb_obj.active
m_row = sheet_obj.max_row

for i in range(1, m_row + 1):
    cell_obj = sheet_obj.cell(row=i, column=1)
    print('-------------------------')
    print('Артикул:' + cell_obj.value)
    url = url_src + cell_obj.value
    r = requests.get(url)
    soup = BeautifulSoup(r.text, 'html.parser')
    url_item = soup.find_all(class_='flex-3')
    url_item_str = str(url_item)
    item = re.findall('href=[\'"]?([^\'" >]+)', url_item_str)
    try:
        item_href = item[0]
    except IndexError:
        print('Артикул на сайте не найден')
        continue

    r2 = r = requests.get('https://www.sds-group.ru' + item_href)
    print('https://www.sds-group.ru' + item_href)
    soup2 = BeautifulSoup(r2.text, 'html.parser')

    try:
        text_item = soup2.find(id="tab-description")
        text = text_item.get_text().replace('  ', '\n')
        print(text)
    except AttributeError:
        print('Ошибка id tab-description не найден')

    try:
        with open('./txt/' + cell_obj.value + '.txt', mode='a', encoding='cp1251') as f:
            text_item = soup2.find(id="tab-description")
            text = text_item.get_text().replace('  ', '\n')
            f.write(text)
    except UnicodeEncodeError:
        print('Проблема с кодировкой')
        print('Текст с заменой:\n' + text)
        text_decode = text.encode('cp1251', errors='ignore').decode('cp1251', errors='ignore')
        with open('./txt/utf-8-encode/' + cell_obj.value + '.txt', mode='a', encoding='cp1251') as file:
            file.write(text_decode)
    except AttributeError:
        print('Ошибка id tab-description не найден')